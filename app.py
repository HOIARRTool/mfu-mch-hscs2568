import re
import html
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import openpyxl


st.set_page_config(
    page_title="HSCS Dashboard",
    page_icon="📊",
    layout="wide"
)

# =========================================================
# Shared thresholds / colors
# =========================================================
H_RED_BG = "#FF2B2B"       # แดงสด
H_ORANGE_BG = "#EF6C00"    # ส้มแก่
H_YELLOW_BG = "#F3E58A"    # เหลืองนวลตา
H_GREEN_BG = "#2E7D32"     # เขียวเข้ม
H_MISSING_BG = "#E8EEF6"   # เทาอ่อนสำหรับช่องไม่มีข้อมูล
H_MISSING_FG = "#64748B"

BASE_DIR = Path(__file__).resolve().parent

HSCS_YEAR_CONFIG = {
    "2568": {
        "label": "ปี 2568",
        "file": BASE_DIR / "HSCS2568_interac.xlsx",
        "sheet": "HSCS2568",
    },
    "2569": {
        "label": "ปี 2569",
        "file": BASE_DIR / "HSCS2569_interac.xlsx",
        "sheet": "HSCS2569",
    },
}

REPORT_URL = "https://sites.google.com/view/mch-hscs67-68/%E0%B8%A0%E0%B8%B2%E0%B8%9E%E0%B8%A3%E0%B8%A7%E0%B8%A1?authuser=0"
REPORT_PREVIEW_IMAGE = BASE_DIR / "hscs_report_preview.png"
MFU_LOGO_URL = "https://mfu.ac.th/fileadmin/_processed_/6/7/csm_logo_mfu_3d_colour_15e5a7a50f.png?raw=true"
HAI_LOGO_URL = "https://github.com/HOIARRTool/appqtbi/blob/main/messageImage_1763018963411.jpg?raw=true"


# =========================================================
# Scoring helpers
# =========================================================
def classify_score(score: float) -> tuple[str, str]:
    """Return status label and color group for a % positive response score."""
    if score < 60:
        return "ควรพัฒนาด่วน", "แดง"
    elif 60 <= score <= 70:
        return "เร่งพัฒนา", "ส้ม"
    elif 70 < score <= 80:
        return "ควรพัฒนาต่อเนื่อง", "เหลือง"
    else:
        return "ควรส่งเสริม", "เขียว"


def heatmap_bg_color(score) -> str:
    if pd.isna(score):
        return H_MISSING_BG
    score = float(score)
    if score < 60:
        return H_RED_BG
    elif 60 <= score <= 70:
        return H_ORANGE_BG
    elif 70 < score <= 80:
        return H_YELLOW_BG
    return H_GREEN_BG


def heatmap_font_color(score) -> str:
    if pd.isna(score):
        return H_MISSING_FG
    score = float(score)
    if score < 60:
        return "#FFFFFF"
    elif 60 <= score <= 70:
        return "#FFFFFF"
    elif 70 < score <= 80:
        return "#111111"
    return "#FFFFFF"


def _score_status(score: float) -> tuple[str, str, str]:
    """Return status label, background color, and text color for a score."""
    if pd.isna(score):
        return "ไม่มีข้อมูล", "#F8FAFC", "#0F172A"
    status, _ = classify_score(float(score))
    bg = heatmap_bg_color(score)
    fg = heatmap_font_color(score)
    return status, bg, fg


def _dimension_sort_key(dim_name: str):
    """Sort dimensions by leading number when available, otherwise by text."""
    m = re.match(r"^\s*(\d+)", str(dim_name))
    if m:
        return (0, int(m.group(1)), str(dim_name))
    return (1, 999, str(dim_name))


def _sub_code_sort_key(code: str):
    """Sort sub codes such as A1, A10, B2 naturally."""
    s = str(code or "")
    m = re.match(r"^([A-Za-z]+)(\d+)$", s)
    if m:
        return (m.group(1), int(m.group(2)))
    return (s, 0)


def dedupe_labels(labels):
    seen = {}
    out = []
    for lab in labels:
        if lab not in seen:
            seen[lab] = 1
            out.append(lab)
        else:
            seen[lab] += 1
            out.append(f"{lab} ({seen[lab]})")
    return out


def get_heatmap_display_mode(unit_count: int) -> dict:
    """
    Control matrix width.

    When many units are displayed, forcing the chart to fit the browser width
    makes each cell too narrow. A fixed wide Plotly canvas keeps numbers legible;
    the user can horizontally scroll / zoom as needed.
    """
    if unit_count <= 1:
        return {"compact": True, "width": 760}
    if unit_count == 2:
        return {"compact": True, "width": 920}
    if unit_count <= 18:
        return {"compact": False, "width": None}

    # Around 40 px per unit keeps the cell text readable in the all-groups view.
    return {"compact": True, "width": max(1450, 220 + unit_count * 42)}


# =========================================================
# Heatmap workbook loader
# =========================================================
def _resolve_header_value(ws, merge_map, row_num, col_num):
    v = ws.cell(row_num, col_num).value
    if v is None and (row_num, col_num) in merge_map:
        v = merge_map[(row_num, col_num)]
    return v


@st.cache_data(show_spinner=False)
def load_heatmap_excel(file_path: Path, sheet_name: str) -> tuple[pd.DataFrame, list[str]]:
    """
    Read HSCS interac workbook.

    Expected sheet structure:
    - Row 1: top group
    - Row 2: division
    - Row 3: unit
    - Column A: dimension
    - Column B: sub-item
    - Columns C onward: scores
    """
    raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    merge_map = {}
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = mr.bounds
        if min_row <= 3:
            top_val = ws.cell(min_row, min_col).value
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    merge_map[(r, c)] = top_val

    data_rows = []
    current_dimension = None

    for r in range(3, len(raw)):  # Excel row 4 onward; pandas is 0-based
        dim = raw.iloc[r, 0] if raw.shape[1] > 0 else None
        sub = raw.iloc[r, 1] if raw.shape[1] > 1 else None

        if pd.notna(dim):
            current_dimension = str(dim).strip()

        numeric_found = False
        for c in range(2, raw.shape[1]):
            val = raw.iloc[r, c]
            if pd.notna(val):
                try:
                    float(val)
                    numeric_found = True
                    break
                except Exception:
                    pass

        if pd.notna(sub) and numeric_found:
            sub_text = str(sub).strip()
            code_match = re.match(r"^([A-Z]\d+)\.\s*", sub_text)
            code = code_match.group(1) if code_match else ""
            full_name = re.sub(r"^[A-Z]\d+\.\s*", "", sub_text).strip()
            data_rows.append((r, {"dimension": current_dimension, "sub_code": code, "sub_name": full_name}))

    if not data_rows:
        raise ValueError("ไม่พบข้อมูล heatmap ในชีตที่เลือก")

    row_indices = [r for r, _ in data_rows]

    score_cols = []
    for c in range(2, raw.shape[1]):
        any_numeric = False
        for r in row_indices:
            val = raw.iloc[r, c]
            if pd.notna(val):
                try:
                    float(val)
                    any_numeric = True
                    break
                except Exception:
                    pass
        if any_numeric:
            score_cols.append(c)

    if not score_cols:
        raise ValueError("ไม่พบคอลัมน์คะแนนในชีตที่เลือก")

    records = []
    groups_found = []

    for r, base in data_rows:
        for c in score_cols:
            col_num = c + 1  # pandas 0-based -> openpyxl 1-based

            top_group = _resolve_header_value(ws, merge_map, 1, col_num)
            division = _resolve_header_value(ws, merge_map, 2, col_num)
            unit = _resolve_header_value(ws, merge_map, 3, col_num)

            top_group = str(top_group).replace("\n", " ").strip() if top_group is not None else ""
            division = str(division).replace("\n", " ").strip() if division is not None else ""
            unit = str(unit).replace("\n", " ").strip() if unit is not None else ""

            if not unit:
                unit = division if division else top_group

            groups_found.append(top_group)

            val = raw.iloc[r, c]
            score = np.nan
            if pd.notna(val):
                try:
                    score = float(val)
                except Exception:
                    score = np.nan

            records.append(
                {
                    "group": top_group,
                    "division": division,
                    "unit": unit,
                    "dimension": base["dimension"],
                    "sub_code": base["sub_code"],
                    "sub_name": base["sub_name"],
                    "score": score,
                    "col_index": c,
                }
            )

    long_df = pd.DataFrame(records)

    ordered_groups = []
    for g in groups_found:
        if g and g not in ordered_groups:
            ordered_groups.append(g)

    return long_df, ordered_groups


def build_overview_df_from_heatmap(long_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build the executive dashboard source from the 'ภาพรวม' column in HSCS*_interac.xlsx.
    If a workbook has no explicit 'ภาพรวม' column, fallback to mean across units.
    """
    df = long_df.copy()

    overall_mask = (
        df["unit"].astype(str).str.strip().eq("ภาพรวม")
        | df["division"].astype(str).str.strip().eq("ภาพรวม")
        | df["group"].astype(str).str.strip().eq("ภาพรวม")
    )

    if overall_mask.any():
        overall_cols = sorted(df.loc[overall_mask, "col_index"].dropna().unique().tolist())
        target_col = overall_cols[0]
        out = df[df["col_index"] == target_col].copy()
        out = out[["dimension", "sub_code", "sub_name", "score"]].rename(columns={"score": "sub_score"})
    else:
        out = (
            df.groupby(["dimension", "sub_code", "sub_name"], dropna=False)["score"]
            .mean()
            .reset_index()
            .rename(columns={"score": "sub_score"})
        )

    out = out.dropna(subset=["sub_score"]).copy()
    out["sub_score"] = pd.to_numeric(out["sub_score"], errors="coerce")
    out = out.dropna(subset=["sub_score"])

    dim_avg = (
        out.groupby("dimension", dropna=False)["sub_score"]
        .mean()
        .rename("dimension_avg")
        .reset_index()
    )
    out = out.merge(dim_avg, on="dimension", how="left")
    out["development_level"] = out["sub_score"].apply(lambda x: classify_score(float(x))[0])
    return out


# =========================================================
# Dashboard overview page
# =========================================================
def _render_dashboard_css():
    st.markdown(
        """
        <style>
        .hscs-hero {
            background: linear-gradient(135deg, #ffffff 0%, #f4f8ff 100%);
            border: 1px solid #dbe5f0;
            border-radius: 22px;
            padding: 18px 22px 18px 24px;
            margin-bottom: 18px;
            box-shadow: 0 8px 24px rgba(15, 23, 42, 0.045);
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 18px;
        }
        .hscs-hero-text { min-width: 0; }
        .hscs-hero h1 {
            color: #173B71;
            margin: 0 0 4px 0;
            font-size: 2.0rem;
            line-height: 1.15;
        }
        .hscs-hero p {
            color: #64748B;
            margin: 0;
            font-size: 1.0rem;
        }
        .hscs-hero-logos {
            display: flex;
            align-items: center;
            justify-content: flex-end;
            gap: 12px;
            flex: 0 0 auto;
        }
        .hscs-hero-logo {
            height: 58px;
            max-width: 155px;
            object-fit: contain;
            background: #FFFFFF;
            border: 1px solid #E2E8F0;
            border-radius: 14px;
            padding: 6px 8px;
            box-shadow: 0 4px 12px rgba(15, 23, 42, 0.08);
        }
        @media (max-width: 760px) {
            .hscs-hero { align-items: flex-start; flex-direction: column; }
            .hscs-hero-logos { justify-content: flex-start; }
            .hscs-hero-logo { height: 48px; max-width: 128px; }
        }
        .hscs-section-title {
            color: #173B71;
            font-weight: 800;
            font-size: 1.35rem;
            margin: 18px 0 10px 0;
            border-left: 5px solid #D7A928;
            padding-left: 12px;
        }
        .hscs-dim-grid {
            display: grid;
            grid-template-columns: repeat(5, minmax(0, 1fr));
            gap: 3px;
            background: #CBD5E1;
            border: 1px solid #CBD5E1;
            border-radius: 14px;
            overflow: hidden;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.055);
            margin-bottom: 16px;
        }
        .hscs-dim-tile {
            min-height: 176px;
            padding: 13px 14px 12px 14px;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
        }
        .hscs-dim-title {
            font-weight: 800;
            font-size: 0.88rem;
            line-height: 1.28;
            min-height: 43px;
            display: -webkit-box;
            -webkit-line-clamp: 2;
            -webkit-box-orient: vertical;
            overflow: hidden;
        }
        .hscs-dim-score {
            text-align: center;
            font-weight: 900;
            font-size: 1.95rem;
            line-height: 1.05;
            margin: 4px 0 2px 0;
        }
        .hscs-dim-status {
            text-align: center;
            font-weight: 700;
            font-size: 0.76rem;
            opacity: 0.92;
            margin-bottom: 5px;
        }
        .hscs-sub-divider {
            height: 1px;
            background: currentColor;
            opacity: 0.42;
            margin: 4px 0 7px 0;
        }
        .hscs-subgrid {
            display: flex;
            flex-wrap: wrap;
            gap: 5px 4px;
            justify-content: center;
        }
        .hscs-subitem {
            min-width: 31%;
            padding: 3px 4px 4px 4px;
            border-radius: 8px;
            text-align: center;
            line-height: 1.08;
            border: 1px solid rgba(255, 255, 255, 0.78);
            box-shadow: inset 0 0 0 1px rgba(0, 0, 0, 0.08), 0 1px 2px rgba(15, 23, 42, 0.14);
        }
        .hscs-subitem span {
            display: block;
            font-weight: 900;
            font-size: 0.70rem;
            text-transform: uppercase;
        }
        .hscs-subitem strong {
            display: block;
            font-weight: 800;
            font-size: 0.70rem;
        }
        .hscs-legend-inline {
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
            align-items: center;
            margin: 10px 0 18px 0;
            color: #334155;
            font-size: 0.84rem;
            font-weight: 700;
        }
        .hscs-legend-dot {
            display: inline-block;
            width: 14px;
            height: 14px;
            border-radius: 4px;
            margin-right: 5px;
            vertical-align: -2px;
        }
        @media (max-width: 1400px) {
            .hscs-dim-grid { grid-template-columns: repeat(3, minmax(0, 1fr)); }
        }
        @media (max-width: 900px) {
            .hscs-dim-grid { grid-template-columns: repeat(1, minmax(0, 1fr)); }
            .hscs-dim-tile { min-height: 150px; }
        }
        .hscs-trend-note {
            color: #64748B;
            font-size: 0.88rem;
            margin: -4px 0 14px 0;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _dimension_key(dim_name: str) -> str:
    """Stable key for comparing dimensions across years; use the leading number when present."""
    m = re.match(r"^\s*(\d+)", str(dim_name or ""))
    return m.group(1) if m else str(dim_name or "").strip()


@st.cache_data(show_spinner=False)
def load_dimension_trend_data() -> tuple[pd.DataFrame, list[str]]:
    """
    Load dimension-level overview scores for every configured year.

    Important: this uses the same workbook loader and the same
    build_overview_df_from_heatmap() logic as the executive dashboard. That keeps
    the trend section aligned with the dashboard's 'ภาพรวม' scores and prevents
    accidental mixing with group-split columns.
    """
    rows = []
    notes = []

    for year, cfg in HSCS_YEAR_CONFIG.items():
        file_path = cfg["file"]
        sheet_name = cfg["sheet"]

        if not file_path.exists():
            notes.append(f"ไม่พบไฟล์ {cfg['label']}: {file_path.name}")
            continue

        try:
            long_df, _ = load_heatmap_excel(file_path, sheet_name=sheet_name)
            overview_df = build_overview_df_from_heatmap(long_df)
        except Exception as exc:
            notes.append(f"โหลดข้อมูล {cfg['label']} ไม่สำเร็จ: {exc}")
            continue

        dim_df = (
            overview_df[["dimension", "dimension_avg"]]
            .drop_duplicates()
            .dropna(subset=["dimension_avg"])
            .copy()
        )
        dim_df["year"] = int(year)
        dim_df["year_label"] = cfg["label"]
        dim_df["dimension_key"] = dim_df["dimension"].map(_dimension_key)

        rows.extend(dim_df.to_dict("records"))

    trend_df = pd.DataFrame(rows)
    if trend_df.empty:
        return trend_df, notes

    # Prefer the newest available dimension label for display, but keep numeric ordering.
    latest_labels = (
        trend_df.sort_values(["dimension_key", "year"])
        .groupby("dimension_key", as_index=False)
        .tail(1)[["dimension_key", "dimension"]]
        .rename(columns={"dimension": "display_dimension"})
    )
    trend_df = trend_df.merge(latest_labels, on="dimension_key", how="left")
    return trend_df, notes


def build_dimension_trend_figure(dim_trend_df: pd.DataFrame, dim_label: str) -> go.Figure:
    """Small per-dimension year trend chart for the dashboard."""
    d = dim_trend_df.sort_values("year").copy()

    y_values = pd.to_numeric(d["dimension_avg"], errors="coerce").dropna().tolist()
    if y_values:
        y_min = max(0, (int(min(y_values) // 10) * 10) - 10)
        y_max = min(100, (int(max(y_values) // 10) * 10) + 20)
        if y_max - y_min < 30:
            y_min = max(0, y_min - 10)
            y_max = min(100, y_max + 10)
    else:
        y_min, y_max = 0, 100

    years = d["year"].astype(int).tolist()
    scores = pd.to_numeric(d["dimension_avg"], errors="coerce").tolist()

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=years,
            y=scores,
            mode="lines+markers+text",
            line=dict(width=2.5, color="#173B71"),
            marker=dict(
                size=11,
                color=[heatmap_bg_color(v) for v in scores],
                line=dict(color="#FFFFFF", width=1.5),
            ),
            text=[f"{v:.1f}%" if pd.notna(v) else "" for v in scores],
            textposition="top center",
            textfont=dict(size=11, color="#0F172A"),
            hovertemplate="ปี %{x}<br>คะแนนเฉลี่ยรายมิติ: %{y:.1f}%<extra></extra>",
            showlegend=False,
        )
    )

    # Show the planning horizon even if future-year data are not available yet.
    # The plotted line still uses only available data points, while the x-axis
    # leaves visual space for future HSCS cycles.
    all_years = [2568, 2569, 2570, 2571, 2572]
    x_min = 2567.75
    x_max = 2572.25

    fig.update_layout(
        title=dict(text=dim_label, font=dict(size=15, color="#34138B"), x=0.0, xanchor="left"),
        paper_bgcolor="#FFFFFF",
        plot_bgcolor="#FFFFFF",
        height=255,
        margin=dict(l=34, r=18, t=58, b=34),
    )
    fig.update_xaxes(
        tickmode="array",
        tickvals=all_years,
        range=[x_min, x_max],
        showgrid=False,
        zeroline=False,
        tickfont=dict(size=11),
    )
    fig.update_yaxes(
        range=[y_min, y_max],
        showgrid=True,
        gridcolor="#E5E7EB",
        zeroline=False,
        tickfont=dict(size=11),
    )
    return fig


def render_dimension_trend_section():
    """Render year-to-year dimension trends, independent of the selected dashboard year."""
    st.markdown('<div class="hscs-section-title">แนวโน้มคะแนนเฉลี่ยรายมิติ</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="hscs-trend-note">เปรียบเทียบคะแนนเฉลี่ยรายมิติจากคอลัมน์ “ภาพรวม” ของแต่ละปี โดยแสดงแกนเวลา 2568–2572 เพื่อใช้ติดตามแนวโน้มระยะต่อไป</div>',
        unsafe_allow_html=True,
    )

    trend_df, notes = load_dimension_trend_data()
    if trend_df.empty:
        st.info("ยังไม่พบข้อมูลเพียงพอสำหรับแสดงแนวโน้มรายมิติ")
        if notes:
            with st.expander("รายละเอียดการโหลดข้อมูลแนวโน้ม", expanded=False):
                for note in notes:
                    st.write(f"- {note}")
        return

    dim_order = (
        trend_df[["dimension_key", "display_dimension"]]
        .drop_duplicates()
        .sort_values("display_dimension", key=lambda s: s.map(_dimension_sort_key))
    )

    cols_per_row = 4
    dims = dim_order.to_dict("records")
    for start in range(0, len(dims), cols_per_row):
        cols = st.columns(cols_per_row)
        for i, dim_info in enumerate(dims[start:start + cols_per_row]):
            dim_key = dim_info["dimension_key"]
            dim_label = dim_info["display_dimension"]
            dim_trend = trend_df[trend_df["dimension_key"] == dim_key].copy()
            fig = build_dimension_trend_figure(dim_trend, dim_label)
            with cols[i]:
                st.plotly_chart(fig, use_container_width=True, key=f"trend_dim_{dim_key}")

    if notes:
        with st.expander("หมายเหตุการโหลดข้อมูลแนวโน้ม", expanded=False):
            for note in notes:
                st.write(f"- {note}")


def render_overview_dashboard_page(heatmap_source: Path, heatmap_sheet: str, year_label: str):
    """Executive dashboard using the selected HSCS interac workbook."""
    _render_dashboard_css()

    long_df, _ = load_heatmap_excel(heatmap_source, sheet_name=heatmap_sheet)
    df = build_overview_df_from_heatmap(long_df)

    overall_score = float(df["sub_score"].mean()) if not df.empty else np.nan
    overall_status, _, _ = _score_status(overall_score)
    urgent_count = int((df["sub_score"] < 60).sum())
    orange_count = int(((df["sub_score"] >= 60) & (df["sub_score"] <= 70)).sum())
    dim_count = int(df["dimension"].nunique())
    sub_count = int(df[["sub_code", "sub_name"]].drop_duplicates().shape[0])

    st.markdown(
        f'<div class="hscs-hero"><div class="hscs-hero-text"><h1>MFU-MCH HSCS Dashboard</h1>'
        f'<p>Hospital Safety Culture Survey: executive overview + drill-down Color-coded Matrix | {html.escape(year_label)}</p></div>'
        f'<div class="hscs-hero-logos"><img class="hscs-hero-logo" src="{MFU_LOGO_URL}" alt="Mae Fah Luang University logo">'
        f'<img class="hscs-hero-logo" src="{HAI_LOGO_URL}" alt="Healthcare Accreditation Institute logo"></div></div>',
        unsafe_allow_html=True,
    )

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Overall Positive Score", f"{overall_score:.1f}%", overall_status)
    m2.metric("จำนวนมิติหลัก", f"{dim_count:,}")
    m3.metric("จำนวนมิติย่อย", f"{sub_count:,}")
    m4.metric("ข้อควรพัฒนาด่วน", f"{urgent_count:,}", f"เร่งพัฒนา {orange_count:,} ข้อ")

    st.markdown('<div class="hscs-section-title">ร้อยละคำตอบเชิงบวก (% Positive Response) จำแนกตามมิติ</div>', unsafe_allow_html=True)

    dim_avg_order = (
        df[["dimension", "dimension_avg"]]
        .drop_duplicates()
        .sort_values("dimension", key=lambda s: s.map(_dimension_sort_key))
    )

    tile_html_parts = []
    for _, dim_row in dim_avg_order.iterrows():
        dim = dim_row["dimension"]
        dim_avg = float(dim_row["dimension_avg"])
        status, bg, fg = _score_status(dim_avg)
        dim_safe = html.escape(str(dim))

        sub_df = df[df["dimension"] == dim].copy()
        sub_df = sub_df.sort_values("sub_code", key=lambda s: s.map(_sub_code_sort_key))

        sub_items = []
        for _, r in sub_df.iterrows():
            code = html.escape(str(r["sub_code"] or ""))
            sub_name = html.escape(str(r["sub_name"] or ""))
            score = float(r["sub_score"])
            sub_status, sub_bg, sub_fg = _score_status(score)
            sub_items.append(
                f'<div class="hscs-subitem" style="background:{sub_bg}; color:{sub_fg};" title="{code}: {sub_name} | {html.escape(sub_status)}">'
                f'<span>{code}</span><strong>{score:.1f}%</strong></div>'
            )

        sub_items_html = "".join(sub_items)
        tile_html_parts.append(
            f'<div class="hscs-dim-tile" style="background:{bg}; color:{fg};" title="{dim_safe}">'
            f'<div class="hscs-dim-title">{dim_safe}</div>'
            f'<div><div class="hscs-dim-score">{dim_avg:.1f}%</div>'
            f'<div class="hscs-dim-status">{html.escape(status)}</div></div>'
            f'<div><div class="hscs-sub-divider"></div>'
            f'<div class="hscs-subgrid">{sub_items_html}</div></div>'
            f'</div>'
        )

    st.markdown(
        f'<div class="hscs-dim-grid">{"".join(tile_html_parts)}</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""
        <div class="hscs-legend-inline">
            <span><i class="hscs-legend-dot" style="background:{H_GREEN_BG};"></i>ควรส่งเสริม &gt; 80</span>
            <span><i class="hscs-legend-dot" style="background:{H_YELLOW_BG};"></i>ควรพัฒนาต่อเนื่อง 70.1–80</span>
            <span><i class="hscs-legend-dot" style="background:{H_ORANGE_BG};"></i>เร่งพัฒนา 60–70</span>
            <span><i class="hscs-legend-dot" style="background:{H_RED_BG};"></i>ควรพัฒนาด่วน &lt; 60</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    render_dimension_trend_section()

    st.markdown('<div class="hscs-section-title">Priority list: ข้อที่มีคะแนนต่ำสุด</div>', unsafe_allow_html=True)
    priority = (
        df.sort_values(["sub_score", "dimension", "sub_code"], ascending=[True, True, True])
        .head(12)
        .rename(
            columns={
                "dimension": "มิติหลัก",
                "sub_code": "รหัส",
                "sub_name": "ชื่อมิติย่อย",
                "sub_score": "% Positive Score",
                "development_level": "ระดับการพัฒนา",
            }
        )
    )
    priority["% Positive Score"] = priority["% Positive Score"].map(lambda x: f"{float(x):.1f}%")
    st.dataframe(
        priority[["มิติหลัก", "รหัส", "ชื่อมิติย่อย", "% Positive Score", "ระดับการพัฒนา"]],
        use_container_width=True,
        hide_index=True,
    )


# =========================================================
# Color-coded Matrix page
# =========================================================
def build_heatmap_figure(long_df: pd.DataFrame, title_text: str = "") -> go.Figure:
    df = long_df.copy()

    row_order = df[["sub_code", "sub_name", "dimension"]].drop_duplicates()
    row_order["row_label"] = row_order["sub_code"].replace("", np.nan).fillna("NA")

    col_order = (
        df[["col_index", "unit", "division", "group"]]
        .drop_duplicates()
        .sort_values("col_index")
        .reset_index(drop=True)
    )
    col_order["col_label"] = dedupe_labels(col_order["unit"].tolist())

    df = df.merge(col_order[["col_index", "col_label"]], on="col_index", how="left")

    row_labels = row_order["row_label"].tolist()
    col_labels = col_order["col_label"].tolist()

    pivot = (
        df.assign(row_label=df["sub_code"].replace("", np.nan).fillna("NA"))
        .pivot_table(index="row_label", columns="col_label", values="score", aggfunc="mean")
        .reindex(index=row_labels, columns=col_labels)
    )

    row_meta = row_order.set_index("row_label")[["sub_code", "sub_name", "dimension"]]
    col_meta = col_order.set_index("col_label")[["unit", "division", "group"]]

    customdata = []
    text_x = []
    text_y = []
    text_values = []
    text_colors = []

    for rlab in pivot.index:
        row_cd = []
        for clab in pivot.columns:
            score = pivot.loc[rlab, clab]
            row_cd.append([
                row_meta.loc[rlab, "sub_code"],
                row_meta.loc[rlab, "sub_name"],
                row_meta.loc[rlab, "dimension"],
                col_meta.loc[clab, "unit"],
                col_meta.loc[clab, "division"],
                col_meta.loc[clab, "group"],
            ])

            if pd.notna(score):
                text_x.append(clab)
                text_y.append(rlab)
                text_values.append(f"{score:.1f}")
                text_colors.append(heatmap_font_color(score))

        customdata.append(row_cd)

    z = pivot.values.astype(float)

    colorscale = [
        [0.0, H_RED_BG], [0.599999, H_RED_BG],
        [0.6, H_ORANGE_BG], [0.7, H_ORANGE_BG],
        [0.700001, H_YELLOW_BG], [0.8, H_YELLOW_BG],
        [0.800001, H_GREEN_BG], [1.0, H_GREEN_BG],
    ]

    fig = go.Figure()

    # Render missing values as a soft grey layer underneath the main heatmap.
    # This prevents blank cells from looking like a display error while keeping
    # them visually distinct from true 0% scores, which remain red.
    missing_mask = np.isnan(z)
    if missing_mask.any():
        missing_z = np.where(missing_mask, 1, np.nan)
        fig.add_trace(
            go.Heatmap(
                z=missing_z,
                x=col_labels,
                y=row_labels,
                zmin=0,
                zmax=1,
                colorscale=[[0, H_MISSING_BG], [1, H_MISSING_BG]],
                showscale=False,
                hoverinfo="skip",
                xgap=1,
                ygap=1,
            )
        )

    fig.add_trace(
        go.Heatmap(
            z=z,
            x=col_labels,
            y=row_labels,
            zmin=0,
            zmax=100,
            colorscale=colorscale,
            showscale=False,
            customdata=customdata,
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "มิติย่อย: %{customdata[1]}<br>"
                "มิติหลัก: %{customdata[2]}<br>"
                "หน่วยงาน: %{customdata[3]}<br>"
                "ฝ่าย/งาน: %{customdata[4]}<br>"
                "กลุ่มงาน: %{customdata[5]}<br>"
                "คะแนน: %{z:.1f}%<extra></extra>"
            ),
            xgap=1,
            ygap=1,
        )
    )

    fig.add_trace(
        go.Scatter(
            x=text_x,
            y=text_y,
            mode="text",
            text=text_values,
            textfont=dict(size=11, color=text_colors),
            hoverinfo="skip",
            showlegend=False,
        )
    )

    # Optional dash marks for cells with no valid denominator / no data.
    missing_x = []
    missing_y = []
    for rlab in pivot.index:
        for clab in pivot.columns:
            if pd.isna(pivot.loc[rlab, clab]):
                missing_x.append(clab)
                missing_y.append(rlab)

    if missing_x:
        fig.add_trace(
            go.Scatter(
                x=missing_x,
                y=missing_y,
                mode="text",
                text=["—"] * len(missing_x),
                textfont=dict(size=11, color=H_MISSING_FG),
                hoverinfo="skip",
                showlegend=False,
            )
        )

    unit_count = len(col_labels)
    display_mode = get_heatmap_display_mode(unit_count)

    fig.update_layout(
        title=None,
        paper_bgcolor="#F8FBFF",
        plot_bgcolor="#F8FBFF",
        margin=dict(l=20, r=20, t=40, b=30),
        height=max(760, 31 * len(row_labels) + 210),
        width=display_mode["width"],
    )

    fig.update_xaxes(title_text="", side="top", tickangle=-35, showgrid=False, tickfont=dict(size=10), automargin=True)
    fig.update_yaxes(title_text="", autorange="reversed", showgrid=False, tickfont=dict(size=11), automargin=True)
    return fig


def render_heatmap_page(heatmap_source: Path, heatmap_sheet: str, selected_page: str, selected_year: str):
    long_df, groups = load_heatmap_excel(heatmap_source, sheet_name=heatmap_sheet)

    if selected_page == "Color-coded Matrix: ภาพรวมทุกกลุ่ม":
        # Use only the workbook columns under group="ภาพรวม".
        # In the rebuilt HSCS*_interac workbook, these columns are aggregated by "งาน"
        # across every กลุ่มตามสรพ. This prevents the all-groups page from showing
        # duplicated unit columns split by group.
        overall_mask = long_df["group"].astype(str).str.strip().eq("ภาพรวม")
        filtered = long_df[overall_mask].copy() if overall_mask.any() else long_df.copy()
        page_title = "Color-coded Matrix: ภาพรวมทุกกลุ่ม"
        page_desc = "Color-coded Matrix ภาพรวมรวมตามงาน ข้ามทุกกลุ่มตาม สรพ."
    else:
        target_group = selected_page.replace("Color-coded Matrix: ", "", 1)
        filtered = long_df[long_df["group"] == target_group].copy()
        page_title = f"Color-coded Matrix: {target_group}"
        page_desc = "Color-coded Matrix แยกตามกลุ่มงานจากแถวบนสุด"

    st.title(page_title)
    st.markdown(f"{page_desc} | ปี {selected_year}")

    if filtered.empty:
        st.warning("ไม่มีข้อมูลสำหรับหน้านี้")
        return

    all_dims = filtered["dimension"].dropna().unique().tolist()
    all_units = filtered["unit"].dropna().unique().tolist()

    with st.sidebar.expander("ตัวกรอง Color-coded Matrix", expanded=True):
        dim_filter = st.multiselect(
            "เลือกมิติหลัก",
            options=all_dims,
            default=all_dims,
            key=f"hm_dim_{selected_year}_{selected_page}",
        )
        unit_filter = st.multiselect(
            "เลือกหน่วยงาน/คอลัมน์",
            options=all_units,
            default=all_units,
            key=f"hm_unit_{selected_year}_{selected_page}",
        )

    filtered = filtered[
        filtered["dimension"].isin(dim_filter) &
        filtered["unit"].isin(unit_filter)
    ].copy()

    if filtered.empty:
        st.warning("ไม่มีข้อมูลหลังจากกรอง")
        return

    c1, c2 = st.columns(2)
    c1.metric("จำนวนมิติย่อย", f"{filtered[['sub_code','sub_name']].drop_duplicates().shape[0]:,}")
    c2.metric("จำนวนหน่วยงาน", f"{filtered['unit'].nunique():,}")

    fig = build_heatmap_figure(filtered, title_text="")
    display_mode = get_heatmap_display_mode(filtered["unit"].nunique())

    if display_mode["compact"] and filtered["unit"].nunique() > 18:
        st.caption("มุมมองนี้มีหลายหน่วยงาน จึงแสดงเป็นแผนภาพกว้างขึ้นเพื่อให้อ่านตัวเลขได้ชัดขึ้น สามารถเลื่อนแนวนอนหรือซูมด้วยเครื่องมือของกราฟได้")

    st.plotly_chart(fig, use_container_width=not display_mode["compact"])

    with st.expander("ดูคำอธิบายรหัสมิติย่อย", expanded=False):
        show_map = (
            filtered[["sub_code", "dimension", "sub_name"]]
            .drop_duplicates()
            .sort_values(["dimension", "sub_code", "sub_name"])
            .rename(columns={"sub_code": "รหัส", "dimension": "มิติหลัก", "sub_name": "ชื่อข้อย่อย"})
        )
        st.dataframe(show_map, use_container_width=True, hide_index=True)


# =========================================================
# Full report page
# =========================================================
def render_full_report_page():
    st.title("📘 การประมวลผล HSCS จากสรพ.")
    st.markdown("Preview card ของแหล่งข้อมูล/เว็บไซต์ต้นทาง")

    card_html = """
    <div style="
        background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%);
        border: 1px solid #dce6f2;
        border-radius: 18px;
        padding: 18px 18px 12px 18px;
        box-shadow: 0 6px 20px rgba(23,59,113,0.06);
        margin-bottom: 14px;">
        <div style="font-size: 24px; font-weight: 700; color: #173B71; margin-bottom: 6px;">
            Hospital Safety Culture Survey (HSCS)
        </div>
        <div style="font-size: 15px; color: #4a678f; line-height: 1.6;">
            การประมวลผล HSCS จาก สรพ. ปี 2567–2568
            หน้านี้ใช้ <b>preview card</b> แทนการฝังเว็บตรง เพื่อให้แสดงผลเสถียรกว่า
        </div>
    </div>
    """
    st.markdown(card_html, unsafe_allow_html=True)

    if REPORT_PREVIEW_IMAGE.exists():
        st.image(str(REPORT_PREVIEW_IMAGE), use_container_width=True)
    else:
        st.info("ไม่พบภาพ preview ในแพ็กเกจ แต่ยังเปิดรายงานฉบับเต็มได้ตามปกติ")

    c1, c2 = st.columns([1, 2])
    with c1:
        st.link_button("🔗 เปิดการประมวลผล HSCS จากสรพ.", REPORT_URL, use_container_width=True)
    with c2:
        st.caption("ถ้าต้องการดูรายละเอียดทั้งหมด แนะนำให้เปิดในแท็บใหม่เพื่อการใช้งานที่ครบถ้วนที่สุด")


# =========================================================
# App shell
# =========================================================
st.sidebar.title("MFU-MCH-HSCS")

selected_year = st.sidebar.selectbox(
    "เลือกปีข้อมูล HSCS",
    options=list(HSCS_YEAR_CONFIG.keys()),
    format_func=lambda y: HSCS_YEAR_CONFIG[y]["label"],
    index=1,
)

selected_config = HSCS_YEAR_CONFIG[selected_year]
heatmap_source = selected_config["file"]
heatmap_sheet = selected_config["sheet"]

if st.sidebar.button("Clear cache / reload data"):
    st.cache_data.clear()
    st.rerun()

if not heatmap_source.exists():
    st.error(
        f"ไม่พบไฟล์ข้อมูล {selected_config['label']}: `{heatmap_source.name}`\n\n"
        "กรุณาวางไฟล์ไว้ในโฟลเดอร์เดียวกับ `app.py` แล้ว deploy ใหม่"
    )
    st.stop()

heatmap_pages = ["Color-coded Matrix: ภาพรวมทุกกลุ่ม"]
try:
    _, group_names = load_heatmap_excel(heatmap_source, sheet_name=heatmap_sheet)
    group_names = [
        g for g in group_names
        if str(g).strip() not in ["", "ภาพรวม", "undefined", "None", "nan"]
    ]
    heatmap_pages += [f"Color-coded Matrix: {g}" for g in group_names]
except Exception as exc:
    st.sidebar.warning(f"โหลดรายชื่อกลุ่มงานไม่ได้: {exc}")

page_options = ["Dashboard ภาพรวม"] + heatmap_pages + ["การประมวลผล HSCS จากสรพ."]

page = st.sidebar.radio(
    "เลือกหน้าที่ต้องการดู",
    page_options,
    index=0,
    key=f"page_{selected_year}",
)

st.sidebar.markdown("---")
st.sidebar.markdown(
    """
**เกณฑ์สีที่ใช้ร่วมกัน**
- 🔴 แดง: % Positive Score < 60 = ควรพัฒนาด่วน
- 🟠 ส้ม: % Positive Score 60–70 = เร่งพัฒนา
- 🟡 เหลือง: % Positive Score 70.1–80 = ควรพัฒนาต่อเนื่อง
- 🟢 เขียว: % Positive Score > 80 = ควรส่งเสริม
- ⚪ เทา: ไม่มีข้อมูล / ไม่มีตัวหารที่ใช้คำนวณ
"""
)

if page == "Dashboard ภาพรวม":
    render_overview_dashboard_page(heatmap_source, heatmap_sheet, selected_config["label"])
elif page == "การประมวลผล HSCS จากสรพ.":
    render_full_report_page()
else:
    render_heatmap_page(heatmap_source, heatmap_sheet, page, selected_year)
